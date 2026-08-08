#!/usr/bin/env python3
"""Render an .xlsx workbook to a simple multi-sheet HTML preview.

Usage:
  python to_html.py <path.xlsx> [output.html]

openpyxl with data_only=True returns None for formula cells that were never
recalculated by Excel/LibreOffice. This script opens the workbook twice:

- formula mode: always available (shows formula text)
- data_only mode: used when cached values exist (after recalc.py)

When a formula cell has no cached value, the HTML shows the formula string
and a workbook-level banner warns that LibreOffice recalc has not run.
"""

from __future__ import annotations

import html
import json
import sys
from pathlib import Path
from typing import Any


def _cell_display(formula_cell: Any, value_cell: Any) -> tuple[str, bool]:
    """Return (display_text, is_unrecalculated_formula)."""
    raw = formula_cell.value
    if raw is None:
        return ("", False)
    if isinstance(raw, str) and raw.startswith("="):
        cached = None if value_cell is None else value_cell.value
        if cached is None:
            return (raw, True)
        return (str(cached), False)
    return (str(raw), False)


def workbook_to_html(xlsx_path: Path, out_path: Path | None = None) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        return {
            "status": "error",
            "error": "openpyxl is required (pip/uv install openpyxl)",
        }

    src = xlsx_path.resolve()
    if not src.exists():
        return {"status": "error", "error": f"File not found: {src}"}

    wb_formulas = load_workbook(src, data_only=False)
    wb_values = load_workbook(src, data_only=True)

    unrecalculated = 0
    sheets_html: list[str] = []

    for sheet_name in wb_formulas.sheetnames:
        ws_f = wb_formulas[sheet_name]
        ws_v = wb_values[sheet_name] if sheet_name in wb_values.sheetnames else None
        rows_html: list[str] = []
        for row_f, row_v in zip(ws_f.iter_rows(), ws_v.iter_rows() if ws_v else []):
            cells: list[str] = []
            for cell_f, cell_v in zip(row_f, row_v if row_v else []):
                text, missing = _cell_display(cell_f, cell_v if ws_v else None)
                if missing:
                    unrecalculated += 1
                    cls = ' class="formula-missing"'
                    title = ' title="Formula not recalculated — run scripts/recalc.py"'
                elif isinstance(cell_f.value, str) and cell_f.value.startswith("="):
                    cls = ' class="formula"'
                    title = f' title="{html.escape(cell_f.value, quote=True)}"'
                else:
                    cls = ""
                    title = ""
                cells.append(f"<td{cls}{title}>{html.escape(text)}</td>")
            if any(c.value is not None for c in row_f):
                rows_html.append("<tr>" + "".join(cells) + "</tr>")
        body = "\n".join(rows_html) if rows_html else "<tr><td><em>(empty sheet)</em></td></tr>"
        sheets_html.append(
            f'<section class="sheet"><h2>{html.escape(sheet_name)}</h2>'
            f'<table>\n{body}\n</table></section>'
        )

    banner = ""
    if unrecalculated:
        banner = (
            '<div class="banner warn">'
            f"⚠️ {unrecalculated} formula cell(s) have no cached values. "
            "Run <code>python scripts/recalc.py &lt;file.xlsx&gt;</code> "
            "(requires LibreOffice headless) before treating numbers as final."
            "</div>"
        )
    else:
        banner = (
            '<div class="banner ok">'
            "Cached values present (workbook appears recalculated)."
            "</div>"
        )

    css = """
    body { font-family: ui-sans-serif, system-ui, sans-serif; margin: 1.5rem; color: #122; background: #f7f6f2; }
    h1 { font-size: 1.25rem; margin-bottom: 0.25rem; }
    .meta { color: #555; font-size: 0.85rem; margin-bottom: 1rem; }
    .banner { padding: 0.75rem 1rem; border-radius: 6px; margin-bottom: 1rem; }
    .banner.warn { background: #fff3cd; border: 1px solid #e0c36a; }
    .banner.ok { background: #e6f4ea; border: 1px solid #8fbf9f; }
    .sheet { margin-bottom: 2rem; overflow-x: auto; }
    table { border-collapse: collapse; background: #fff; font-size: 0.85rem; }
    td { border: 1px solid #ddd; padding: 0.25rem 0.45rem; white-space: nowrap; max-width: 28rem; overflow: hidden; text-overflow: ellipsis; }
    td.formula { color: #111; }
    td.formula-missing { color: #8a4b00; background: #fff8e8; font-family: ui-monospace, monospace; }
    """

    doc = f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<title>{html.escape(src.name)} — preview</title>
<style>{css}</style>
</head>
<body>
<h1>{html.escape(src.name)}</h1>
<p class="meta">HTML preview generated from openpyxl. Not a live spreadsheet.</p>
{banner}
{"".join(sheets_html)}
</body>
</html>
"""

    target = out_path.resolve() if out_path else src.with_suffix(".html")
    target.write_text(doc, encoding="utf-8")
    return {
        "status": "success",
        "file": str(src),
        "html": str(target),
        "unrecalculated_formula_cells": unrecalculated,
    }


def main() -> None:
    if len(sys.argv) < 2:
        print("Usage: python to_html.py <path.xlsx> [output.html]", file=sys.stderr)
        sys.exit(2)
    xlsx = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else None
    result = workbook_to_html(xlsx, out)
    print(json.dumps(result, indent=2))
    sys.exit(0 if result.get("status") == "success" else 1)


if __name__ == "__main__":
    main()
